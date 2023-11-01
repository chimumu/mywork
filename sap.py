import  pymssql
from time import sleep
from openpyxl import Workbook
from openpyxl import load_workbook
import  os

def  r_excel():
    conn = pymssql.connect('10.0.3.13', 'muqi', 'Hippih01', 'mos')
    cursor_1 = conn.cursor()
    directory='D:\\app\\技术资料\\新建文件夹\\账户周一\\'
    file_names = os.listdir(directory)

    # 遍历文件名字列表
    for file_name in file_names:
        # 打印文件名字file_name.strip('.')
        cc=file_name.split('.')
    
    #遍历excel
        wk=load_workbook('D:\\app\\技术资料\\新建文件夹\\账户周一\\'+str(cc[0])+'.'+'xlsx')
        sheet=wk['Sheet1']
        maxrow=sheet.max_row
        maxcou=sheet.max_column
        list2=[]
        for i in range(1,maxrow+1):
            list1=[]
            for j in range(1,9):
                if sheet.cell(i,j).value is None:
                   list1.append('NO')
                else:
                   list1.append(sheet.cell(i,j).value)
            list1.append(cc[0])
            list2.append(tuple(list1))
            del list1
        sql1='INSERT INTO SapAccount (username,fullname,group1,account_no,type,Department,function1,Room,CreateTime) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)'
        cursor_1.executemany(sql1,list2)
        conn.commit()
    
            
            

def sqlserver():
    # 可以进入 connect 方法里面查看更多参数
    conn = pymssql.connect('10.0.3.13','muqi','Hippih01','mos')

    # 游标使用注意事项
    # 一个连接一次只能有一个游标的查询处于活跃状态，如下：
    cursor_1 = conn.cursor()
    cursor_1.execute('drop table SapAccount')
    cursor_1.execute('''CREATE TABLE SapAccount(
                     username varchar(50),
                     fullname  varchar(50),
                     group1   varchar(50),
                     account_no  varchar(50),
                     type    varchar(50),
                     Department  varchar(50),
                     function1    varchar(50),
                     Room       varchar(50),
                     CreateTime   varchar(50) )''')
    conn.commit()
#sqlserver()
r_excel()