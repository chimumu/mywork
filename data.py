from openpyxl import Workbook
from openpyxl import load_workbook
import openpyxl
from openpyxl.styles import Font  # 导入字体模块
from openpyxl.styles import PatternFill  # 导入填充模块
import sys

import datetime
import os

# 获得当前时间
def get_dir():
    list_dir=[]
    with open(r'D:\\ll\test\\test1.txt', 'w') as f2:

      f2.truncate()
    with open(r'D:\\ll\test\\test.txt','r') as f:
         for i in f.readlines():
             #a=i.strip(' ')
             with open(r'D:\\ll\test\\test1.txt','a') as f1:
                 f1.write(str(i.strip( ))+'\n')
    with open(r'D:\\ll\test\\test1.txt','r') as f3:
                 for j in f3.readlines():
                     a=os.path.basename(j)
                     aaa=a.strip('\n')
                     aaaa=aaa.strip('\.xlsx')
                     list_dir.append(aaaa)
    return list_dir
def get_data(dir1,dir2):
    now = datetime.datetime.now()
    # 转换为指定的格式
    otherStyleTime = now.strftime("%Y%m%d%H%M")
    workbook1 = Workbook(dir2)
    workbook1.save(dir2)
    workbook = load_workbook(dir1)
    workbook2 = load_workbook(dir2)
    sheet = workbook['Sheet1']
    sheet1 = workbook2['Sheet']

    list1 = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']
    list2 = ['Initial(mAh)', '1C容量(mAh)', '1C+0.2C容量(mAh)', 'Vavg1C(V)', '2C容量(mAh)', '2C+0.2C容量(mAh)', 'Vavg2C(V)',
             '直流内阻(mΩ)']
    list3 = ['455-630', '≥630', '≥700', '≥1.197', '≥601', '≥700', '≥1.146', '≤80']
    lists=['Cinitial','C1C','C0.2C-A','Vavg1C','C2C','C0.2C-B','Vavg2C','RDC']
    sheet1['A3'] = '范围'
    list4 = []  # 2规格
    list5 = []  # 14
    list6 = []  # 16
    list7 = []  # 20
    list8 = []  # 22
    list14 = []
    list15 = []
    list_dir=get_dir()
    ws1=sheet.max_row
    for j in range(len(list1)):
        sheet1[list1[j] + '1'] = lists[j]
        sheet1[list1[j] + '2'] = list2[j]
        sheet1[list1[j] + '3'] = list3[j]
    for i in range(5, ws1, 5):  # 220
        value = sheet['C' + str(i)].value
        list15.append(value)
    for i in range(3, ws1, 5):  # 140
        value = sheet['C' + str(i)].value
        list14.append(value)
    for i in range(2, ws1, 5):  # 2
        value = sheet['B' + str(i)].value
        list4.append(value)
    for i in range(3, ws1, 5):  # 14
        value = sheet['B' + str(i)].value
        list5.append(value)
    for i in range(4, ws1, 5):  # 16
        value = sheet['B' + str(i)].value
        list6.append(value)
    for i in range(5, ws1, 5):  # 20
        value = sheet['B' + str(i)].value
        list7.append(value)
    for i in range(6, ws1+1, 5):  # 22
        value = sheet['B' + str(i)].value
        list8.append(value)

    
    for q in range(0, len(list5)):
        ss = q + 4
        sheet1['A' + str(ss)] = list_dir[q]
        fille = PatternFill('solid', fgColor="FF00FF")
        if 455 < int(list4[q]) < 630:  # 数据1
            sheet1.cell(ss, 2, list4[q])
        else:
            sheet1.cell(ss, 2, list4[q]).fill = fille
        if int(list5[q]) >= 630:  # 数据2
            sheet1['C' + str(ss)]= list5[q]
        else:
            sheet1.cell(ss, 3, list5[q]).fill = fille

        if (int(list5[q]) + int(list6[q])) >= 700:  # 数据3
            sheet1['D' + str(ss)] = list6[q] + list5[q]
        else:
            sheet1.cell(ss, 4, int(list5[q]) + int(list6[q])).fill = fille

        if list14[q]!= '0' and ((float(list14[q]) * 1000) / float(list5[q])) >= 1.197:  # 数据4
            sheet1['E' + str(ss)] = round((float(list14[q]) * 1000) / float(list5[q]),3)
        elif list14[q] == '0':
            sheet1.cell(ss, 5, 0).fill = fille
        else:
            sheet1.cell(ss, 5, round((float(list14[q]) * 1000) / float(list5[q]),3)).fill = fille
        if int(list7[q]) >= 600:  # 数据5
            sheet1['F' + str(ss)] = list5[q]
        else:
            sheet1.cell(ss, 6, list5[q]).fill = fille
        if (int(list7[q]) + int(list8[q])) >= 700:  # 数据6
            sheet1['G' + str(ss)] = list7[q] + list8[q]
        else:
            sheet1.cell(ss, 7, int(list7[q]) + int(list8[q])).fill = fille

        if list15[q] != '0' and ((float(list15[q]) * 1000) / float(list7[q])) >= 1.146:  # 数据7
            sheet1['H' + str(ss)] = round((float(list15[q]) * 1000) / float(list7[q]),3)
        elif list15[q] == '0':
            sheet1.cell(ss, 8, 0).fill = fille
        else:
            sheet1.cell(ss, 8, round((float(list15[q]) * 1000) / float(list7[q]),3)).fill = fille

    for kv in range(4, sheet1.max_row + 1):
        fille1 = PatternFill('solid', fgColor="FF00FF")
        a1 = float(sheet1.cell(kv, 5).value) - float(sheet1.cell(kv, 8).value)
        if  a1!= 0 and (float(a1) / 0.7) * 1000 <= 80:
            sheet1.cell(kv, 9, round((float(a1) / 0.7) * 1000,1))
        elif float(a1) == 0:
            sheet1.cell(kv,9,0).fill = fille1
        else:
            sheet1.cell(kv, 9, round((float(a1) / 0.7) * 1000),1).fill = fille1
    fille = PatternFill('solid', fgColor="FF00FF")
    sheet1.cell(sheet1.max_row+1,1,'数据不合格色').fill=fille
    sheet1.column_dimensions['A'].width = 25.0
    sheet1.column_dimensions['B'].width = 15.0
    sheet1.column_dimensions['C'].width = 15.0
    sheet1.column_dimensions['D'].width = 15.0
    sheet1.column_dimensions['E'].width = 15.0
    sheet1.column_dimensions['F'].width = 15.0
    sheet1.column_dimensions['G'].width = 15.0
    sheet1.column_dimensions['H'].width = 15.0
    sheet1.column_dimensions['I'].width = 15.0
    workbook2.save(dir2)
      

if __name__=='__main__':
  
  now = datetime.datetime.now()
    # 转换为指定的格式
  otherStyleTime = now.strftime("%Y%m%d%H%M")
  ll='D:\\ll\\test\\'+str(otherStyleTime)+'.'+'xlsx'

  get_data('D:\\ll\\test\\data.xlsx',ll)
  #get_dir()

# for row in sheet.iter_rows(min_row=1, max_row=10):
#    for cell in row:
#        print(cell.value)
# workbook1.save('D:\\ll\\电池8.30\\电池8.30\\qa1.xlsx')