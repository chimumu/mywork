from elasticsearch import Elasticsearch
import re
import smtplib
import  pymssql
from email.header import Header
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
from selenium import webdriver
from time import sleep
from selenium.webdriver.common.by import By
from openpyxl import Workbook
from openpyxl import load_workbook
import ast
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
#from  PIL import Image
#import  numpy as np



def es1():
    es = Elasticsearch("http://10.0.3.116:9200/", http_auth=('elastic', '123456'))
    body={
              "query": {
                "bool": {
                  "must": [
                    {"range": {
                      "@timestamp": {
                        "gte": "now-15m",
                        "lte": "now"
                      }
                    }},
                    {"regexp": {
                      "sc-status": "[3-5].*"
                    }}
                  ]
                }
              },
              "size": 0,
              "aggs": {
                "brandAgg": {
                  "terms": {
                    "field": "cs-username.keyword",
                    "order": {
                      "_count": "desc"
                    },
                    "size": 30
                  }
                }
              }
           }
    result1=es.search(index='logstash-ex*', body=body)

    dict={}
    cc=str(result1).split("{")
    for k in cc[7:]:
        p=(k.split(","))
        names=p[0].split(':')[1].strip('\' \'')
        xx=repr(names).split('\\\\')
        if len(xx)>1:
            nx=xx[0].strip('\' \'')
            nz=xx[2].strip('\' \'')
            ns=nx+'\\'+nz
            nx=ns.split('\' \'')
            str1 = re.compile('[0-9]+')
            str2 = re.search(str1, p[1])
            str3 = str2.group(0)
            dict[nx[0]] = str3
        else:
            str1 = re.compile('[0-9]+')
            str2 = re.search(str1, p[1])
            str3 = str2.group(0)
            dict[names]=str3
    names1=[]
    sums=[]
    for j in  dict:
      if j != '-':
        body2 = {
            "query": {
                "bool": {
                    "must": [
                        {"range": {
                            "@timestamp": {
                                "gte": "now-15m",
                                "lte": "now"
                            }
                        }},
                        {"term": {
                            "cs-username.keyword": {
                                "value":  j
                            }
                        }}
                    ]
                }
            }
        }
        result2 = es.search(index='logstash-ex*', body=body2)
        cc1 = str(result2).split("{")
        res1 = re.compile('[0-9]+')
        res2 = re.search(res1, cc1[4])
        values = res2.group(0)
        if int(dict[j])/int(values)*100 == 100 :
              names1.append(j)
              sums.append(values)
        #print('15min用户失败率100%\n用户名:{}失败率:{}'.format(j,int(dict[j])/int(values)*100),)
        with open("C://ll//失败用户.txt",'w+') as f:
            for i in range(len(names1)):
                if int(sums[i]) >= 3000:
                    f.write(names1[i]+'   失败率100%  '+'    失败次数：' +sums[i]+'\n')

    for zz in sums:
        if int(zz) >=3000:
           sendMail()
           exit()


def es():#7d
    conn = pymssql.connect('10.0.3.13', 'muqi', 'Hippih01', 'mos')
    es = Elasticsearch("http://10.0.3.118:9200/",http_auth=('elastic', '123456'))
    body={
        "query": {
            "bool": {
                "must": [
                    {"range": {
                        "@timestamp": {
                            "gte": "now-7d",
                            "lte": "now"
                        }
                    }},
                    {
                        "term": {
                            "fields.type.keyword": {
                                "value": "SCMCENTER"
                            }
                        }
                    }
                ]
            }
        },
        "size": 100000,
        "_source": ["Operator"]
        }
    result = es.search(index='basicaudit*', body=body)
    cc = str(result).split("{")
    list1=[]
    list2=[]
    list3=[]
    list4=[]
    list5=[]
    list6=[]
    list7=[]
    list8=[]
    count1={}
    for i in cc[5::]:
        a=i.split(',')
        list1.append(a[0])
    for j in range(1,len(list1),2):
        aa=list1[j].strip('}').strip(']').strip('}')
        cc=aa.split(':')
        if len(cc) != 1:
         bb=cc[1].strip('\'\'').strip(' ').strip('\'')
         list2.append(bb)
    count = {}
    for char in list2:
        if char in count:
            count[char] += 1
        else:
            count[char] = 1
    for key, value in excel().items():
       for key1,value1 in count.items():
           if key == key1:
              list3.append(value)
              list6.append(key1)
              list7.append(value)
              list8.append(value1)
             #print(key1,value,value1)
    print(len(list1))
    for w in list3:
        if w in count1:
            count1[w]+=1
        else:
            count1[w]=1
    workbook1 = Workbook('D:\\ll\\gantt_project\\test.xlsx')
    sheet = workbook1.create_sheet('7dsite',index=0)
    sheet = workbook1.create_sheet('7duser',index=0)
    workbook1.save('D:\\ll\\gantt_project\\test.xlsx')
    workbook2=load_workbook('D:\\ll\\gantt_project\\test.xlsx')
    sheet1=workbook2['7dsite']
    sheet2 = workbook2['7duser']

    for h in count1:
        list4.append(h)
    for h1 in count1.values():
        list5.append(h1)
    for h2 in range(2,len(list5)+2):
        sheet1.cell(h2, 2).value = list5[h2-2]
    for h3 in range(2,len(list4)+2):
        sheet1.cell(h3, 1).value = list4[h3-2]
    for h4 in range(2,len(list6)+2):
        sheet2.cell(h4, 1).value = list6[h4-2]
    for h5 in range(2,len(list7)+2):
        sheet2.cell(h5, 2).value = list7[h5-2]
    for h6 in range(2,len(list8)+2):
        sheet2.cell(h6, 3).value = list8[h6-2]

    cursor_1 = conn.cursor()
    # cursor_1.execute('''insert into employees (id,name,age) values(1,'muqi',12)''')
    list9=[]
    for nu in range(0,len(list6)):
      tup=(list6[nu],list7[nu],list8[nu])
      list9.append(tup)
      del  tup
      #sql1='insert into IOMS (name,site,count) values('+str(list6[nu])+','+str(list7[nu])+','+str(list8[nu])+')'
    sql1 = "INSERT INTO IMOS (name,site,count) VALUES (%s,%s,%s)"
    cursor_1.executemany(sql1,list9)
    conn.commit()
##########设置表头###################
    column_titles = ["USER", "SITE", "PV"]
    # 写入列标题到Excel的第一行
    for col_num, title in enumerate(column_titles, start=1):
        cell = sheet2.cell(row=1, column=col_num)
        cell.value = title
    workbook2.save('D:\\ll\\gantt_project\\test.xlsx')

    column_titles1 = ["SITE", "数量"]

    # 写入列标题到Excel的第一行
    for col_num1, title1 in enumerate(column_titles1, start=1):
        cell = sheet1.cell(row=1, column=col_num1)
        cell.value = title1

    workbook2.save('D:\\ll\\gantt_project\\test.xlsx')
def es2():#30d
    conn = pymssql.connect('10.0.3.13', 'muqi', 'Hippih01', 'mos')
    es = Elasticsearch("http://10.0.3.116:9200/", http_auth=('elastic', '123456'))
    body = {
        "query": {
            "bool": {
                "must": [
                    {"range": {
                        "@timestamp": {
                            "gte": "now-30d",
                            "lte": "now"
                        }
                    }},
                    {
                        "term": {
                            "fields.type.keyword": {
                                "value": "SCMCENTER"
                            }
                        }
                    }
                ]
            }
        },
        "size": 50000,
        "_source": ["Operator"]
    }
    result = es.search(index='basicaudit*', body=body)
    cc = str(result).split("{")
    list1 = []
    list2 = []
    list3 = []
    list4 = []
    list5 = []
    list6 = []
    list7 = []
    list8 = []
    list10=[]

    count1 = {}
    for i in cc[5::]:
        a = i.split(',')
        list1.append(a[0])
    for j in range(1, len(list1),2):
        aa = list1[j].strip('}').strip(']').strip('}')
        cc = aa.split(':')
        if len(cc) != 1:
            bb = cc[1].strip('\'\'').strip(' ').strip('\'')
            list2.append(bb)

    count = {}
   #时间
    for  h in range(0, len(list1),2):
        aa = list1[h]
        cc = aa.split(':')
        
        if len(cc) != 1:
            bb = cc[1].strip('\'\'').strip(' ').strip('\'')
            list10.append(bb)


    cursor_1 = conn.cursor()
    for bb in range(0,len(list2)):

                list14=[]
                qq=list10[bb].split('-')

                list13=(qq[1],list2[bb])
                list14.append(list13)
                sql2 = "INSERT INTO TIME3 (time1,name) VALUES (%s,%s)"
                cursor_1.executemany(sql2, list14)
                conn.commit()
                del list13

    list15=[]
  


              
    
    for char in list2:
        if char in count:
            count[char] += 1
        else:
            count[char] = 1
    for key, value in excel().items():
        for key1, value1 in count.items():
            if key == key1:
                list3.append(value)
                list6.append(key1)
                list7.append(value)
                list8.append(value1)
                #print(key1, value, value1)

    for w in list3:
        if w in count1:
            count1[w] += 1
        else:
            count1[w] = 1
    #workbook1 = Workbook('D:\\ll\\gantt_project\\test.xlsx')
    workbook2 = load_workbook('D:\\ll\\gantt_project\\test.xlsx')
    sheet1 = workbook2.create_sheet('30dsite',index=0)
    sheet2 = workbook2.create_sheet('30duser',index=0)
    #workbook1.save('D:\\ll\\gantt_project\\test.xlsx')

    #sheet1 = workbook2['30dsite']
    #sheet2 = workbook2['30duser']

    for h in count1:
        list4.append(h)
    for h1 in count1.values():
        list5.append(h1)
    for h2 in range(2, len(list5)+2):
        sheet1.cell(h2, 2).value = list5[h2-2]
    for h3 in range(2, len(list4)+2):
        sheet1.cell(h3, 1).value = list4[h3-2]
    for h4 in range(2, len(list6)+2):
        sheet2.cell(h4, 1).value = list6[h4-2]
    for h5 in range(2, len(list7)+2):
        sheet2.cell(h5, 2).value = list7[h5-2]
    for h6 in range(2, len(list8)+2):
        sheet2.cell(h6, 3).value = list8[h6-2]


    # cursor_1.execute('''insert into employees (id,name,age) values(1,'muqi',12)''')
    list9 = []
    for nu in range(0, len(list6)):
        tup = (list6[nu], list7[nu], list8[nu])
        list9.append(tup)
        del tup
        # sql1='insert into IOMS (name,site,count) values('+str(list6[nu])+','+str(list7[nu])+','+str(list8[nu])+')'


    sql1 = "INSERT INTO IMOS3 (name,site,count) VALUES (%s,%s,%s)"

    cursor_1.executemany(sql1,list9)
    #
    conn.commit()

    ##########设置表头###################
    column_titles = ["USER", "SITE", "PV"]

    # 写入列标题到Excel的第一行
    for col_num, title in enumerate(column_titles, start=1):
        cell = sheet2.cell(row=1, column=col_num)
        cell.value = title
    workbook2.save('D:\\ll\\gantt_project\\test.xlsx')

    column_titles1 = ["SITE", "数量"]

    # 写入列标题到Excel的第一行
    for col_num1, title1 in enumerate(column_titles1, start=1):
        cell = sheet1.cell(row=1, column=col_num1)
        cell.value = title1
    workbook2.save('D:\\ll\\gantt_project\\test.xlsx')


def sendMail():
    # 定义相关数据,请更换自己的真实数据
    smtpserver = 'mail.hi-p.com'
    sender = 'elk@hi-p.com'
    #receiver可设置多个，使用“,”分隔
    receiver = 'qi.mu@hi-p.com,Qiao.JianQing@hi-p.com,Bo.Chen@hi-p.com'

    msg = MIMEMultipart()
    boby = """
        <h3>elk邮箱登录失败率检测15min/per</h3>
        <h4>【注】15min 失败率100%的用户,超过3000的用户见附件<h4>
        <p>状态如下：</p>
        <p>
        <span>1.访问状态码百分比：<br><img src="cid:image0"></br></span>
        <span>2.TOP10访问ip:<br><img src="cid:image1"></br> </span>
        <span>3.TOP10失败用户<br><img src="cid:image2"></br></span>
        <span>4.IP来源地百分比<br><img src="cid:image3"></br></span>
        </p>
        <span>url:<p>http://10.0.3.118:5601/app/dashboards#/view/b7f9d7f0-08d3-11ee-b540-ffe6baaaa9b7?_g=h@c823129</p></span>
        <p>
    """
    list1=['status.png','topip.png','user.png','ipsource.png']
    for i in range(0,len(list1)):
       fp = open(list1[i], 'rb')
       images = MIMEImage(fp.read())
       fp.close()
       images.add_header('Content-ID', '<image'+str(i)+'>')
       msg.attach(images)
    att1 = MIMEText(open('C://ll//失败用户.txt', 'rb').read(), 'base64', 'utf-8')
    att1["Content-Type"] = 'application/octet-stream'
    # 这里的filename可以任意写，写什么名字，邮件中显示什么名字
    att1["Content-Disposition"] = 'attachment; filename="faileduser.txt"'
    msg.attach(att1)
    mail_body = MIMEText(boby, _subtype='html', _charset='utf-8')
    msg['Subject'] = Header("elk检测报告", 'utf-8')
    msg['From'] = sender
    receivers = receiver
    toclause = receivers.split(',')
    msg['To'] = ",".join(toclause)
    print(msg['To'])
    msg.attach(mail_body)
    # 登陆并发送邮件
    try:
        smtp = smtplib.SMTP()
        ##打开调试模式
        # smtp.set_debuglevel(1)
        smtp.connect(smtpserver)
        smtp.sendmail(sender, toclause, msg.as_string())
    except:
        print("邮件发送失败！！")
    else:
        print("邮件发送成功")
    finally:
        smtp.quit()
def  excel():
    list1=[]
    list2=[]
    dict1={}
    workbook2 = load_workbook('D:\\ll\\gantt_project\\User.xlsx')
    sheet1 = workbook2['Sheet2']
    for i in range(1,sheet1.max_row+1):
        list1.append(sheet1.cell(i,1).value)
    for j in range(1,sheet1.max_row+1):
        list2.append(sheet1.cell(j,3).value)
    for k in range(len(list1)):
        dict1[list1[k]]=list2[k]
    return dict1

def sqlserver1():
     conn=pymssql.connect('10.0.3.13','muqi','Hippih01','mos')
     cs1=conn.cursor()
     cs1.execute('drop table SITE_IMOS')
     cs1.execute('''CREATE TABLE SITE_IMOS
                (name varchar(50),
               site  varchar(50))''')
     ws=load_workbook('D:\\ll\\gantt_project\\site_imos.xlsx')
     sheet=ws['Sheet1']
     dirc={}
     list2=[]
     for i in range(1,sheet.max_row+1):
         list1=[]
         for j in range(1,sheet.max_column):
             list1.append(sheet.cell(i,j).value)
             dirc[sheet.cell(i,j).value]=sheet.cell(i,j+1).value
             list1.append(sheet.cell(i, j+1).value)
         list2.append(tuple(list1))
         del list1
     del dirc['Account']
     print(list2)
     sql1='INSERT INTO SITE_IMOS(name,site) VALUES (%s,%s)'
     cs1.executemany(sql1,list2)
     conn.commit()
    
def  sqlserver():
    # 可以进入 connect 方法里面查看更多参数
    conn = pymssql.connect('10.0.3.13','muqi','Hippih01','mos')

    # 游标使用注意事项
    # 一个连接一次只能有一个游标的查询处于活跃状态，如下：
    cursor_1 = conn.cursor()
    cursor_1.execute('drop table TIME3')
    cursor_1.execute('drop table IMOS3')
    cursor_1.execute('drop table IMOS')
    cursor_1.execute('''CREATE TABLE IMOS
                 (name varchar(50),
                  site  varchar(50),
    			  count    int )''')
    cursor_1.execute('''CREATE TABLE IMOS3
             (name varchar(50),
              site  varchar(50),
			  count    int )''')
    cursor_1.execute('''CREATE TABLE TIME3
                 (time1 varchar(50),
                  name   varchar(50))''')
    #cursor_1.execute('''insert into employees (id,name,age) values(1,'muqi',12)''')
    conn.commit()
sqlserver1()
sqlserver()
es()
es2()
